import requests
import pandas as pd
import time

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ==============================
# CONFIG
# ==============================

BASE_URL = "https://phagesdb.org/api/phages/"
PAGE_SIZE = 100
MAX_PAGES = 2

OUTPUT_FILE = "Phage_Statistical_Dataset.xlsx"

# ==============================
# FETCH DATA
# ==============================

def fetch_first_pages():

    phages = []

    for page in range(1, MAX_PAGES + 1):

        print(f"Downloading page {page}...")

        params = {
            "page": page,
            "page_size": PAGE_SIZE
        }

        response = requests.get(BASE_URL, params=params)

        if response.status_code != 200:
            print("API Error:", response.status_code)
            break

        data = response.json()

        results = data["results"]

        if not results:
            break

        phages.extend(results)

        time.sleep(0.2)

    return phages


# ==============================
# EXTRACT DATA
# ==============================

def extract_data(phages):

    records = []

    for p in phages:

        record = {

            "Phage Name": p.get("phage_name"),

            "Country": p.get("found_country"),
            "City": p.get("found_city"),
            "Year": p.get("found_year"),

            "Morphotype": p.get("morphotype"),

            "Genome Length (bp)": p.get("genome_length"),
            "GC Percent": p.get("gcpercent"),

            "Number ORFs": p.get("num_ORFs"),
            "Number tRNAs": p.get("num_tRNAs"),

            "Cluster": (
                p.get("pcluster", {})
                .get("cluster")
                if p.get("pcluster")
                else None
            ),

            "Subcluster": (
                p.get("psubcluster", {})
                .get("subcluster")
                if p.get("psubcluster")
                else None
            ),

            "Host Genus": (
                p.get("host_genus", {})
                .get("genus_name")
                if p.get("host_genus")
                else None
            ),

            "Host Species": (
                p.get("host_species", {})
                .get("species_name")
                if p.get("host_species")
                else None
            )

        }

        records.append(record)

    return pd.DataFrame(records)


# ==============================
# FORMAT EXCEL
# ==============================

def format_excel(file):

    wb = load_workbook(file)

    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for ws in wb.worksheets:

        # Freeze first row
        ws.freeze_panes = "A2"

        # Add filters
        ws.auto_filter.ref = ws.dimensions

        for col in ws.columns:

            max_length = 0
            column = col[0].column

            letter = get_column_letter(column)

            # Header formatting
            cell = ws[f"{letter}1"]

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment

            # Auto width
            for c in col:

                try:
                    if c.value:
                        max_length = max(
                            max_length,
                            len(str(c.value))
                        )
                except:
                    pass

            adjusted_width = max_length + 2

            ws.column_dimensions[letter].width = adjusted_width

    wb.save(file)


# ==============================
# SAVE MULTIPLE SHEETS
# ==============================

def save_to_excel(df):

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        # Main dataset
        df.to_excel(
            writer,
            sheet_name="All_Phages",
            index=False
        )

        # By Host Genus
        genus_group = (
            df.groupby("Host Genus")
            .size()
            .reset_index(name="Count")
        )

        genus_group.to_excel(
            writer,
            sheet_name="Host_Genus_Count",
            index=False
        )

        # By Cluster
        cluster_group = (
            df.groupby("Cluster")
            .size()
            .reset_index(name="Count")
        )

        cluster_group.to_excel(
            writer,
            sheet_name="Cluster_Count",
            index=False
        )

    format_excel(OUTPUT_FILE)

    print("\nExcel file created:")
    print(OUTPUT_FILE)


# ==============================
# MAIN
# ==============================

def main():

    print("Downloading data...\n")

    phages = fetch_first_pages()

    print(f"\nTotal phages: {len(phages)}")

    df = extract_data(phages)

    print("Saving Excel...")

    save_to_excel(df)


if __name__ == "__main__":
    main()